"""
Сборка итогового .xlsx строго по визуальному формату эталона (см.
docs/ПРАВИЛА_ОБРАБОТКИ.md, раздел 3) на основе сетки заданного модуля-заказа.

Запуск: python3 src/build_xlsx.py [модуль_заказа] [путь_к_файлу.xlsx]
По умолчанию модуль - order_grid (заказ на 22.08.2026).
"""
import importlib
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"

# --- стили, снятые с эталонного файла -------------------------------------
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
FILL_PREMIUM_HEADER = PatternFill("solid", fgColor="CCCCCC")
FILL_ROW3_BLUE = PatternFill("solid", fgColor="D0E0E3")
FILL_MEAL_PINK = PatternFill("solid", fgColor="EAD1DC")
FILL_BASE_GREY = PatternFill("solid", fgColor="D9D9D9")

THIN = Side(style="thin")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FONT_TITLE = Font(name="Arial", size=24, bold=True)
FONT_HEADER = Font(name="Nunito", size=12, bold=True)
FONT_ROW3 = Font(name="Nunito", size=16, bold=True)
FONT_ROW4 = Font(name="Nunito", size=16, bold=True)
FONT_MEAL = Font(name="Nunito", size=16, bold=True)
FONT_DISH_NAME = Font(name="Nunito", size=14, bold=True)
FONT_DISH_VAL = Font(name="Nunito", size=16, bold=True)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Ширины столбцов - сняты с эталонного файла (docs/ПРАВИЛА_ОБРАБОТКИ.md, раздел 3):
# A..E - индивидуальные, F и G вместе на 14.63, все столбцы данных дальше -
# на "ширине по умолчанию листа" эталона (12.6328125), которая шире обычного
# дефолта Excel (~8.43) - без неё колонки выглядят у́же, чем в примере.
COL_WIDTHS = {"A": 4.75, "B": 54.88, "C": 22.88, "D": 15.88, "E": 17.63, "F": 14.63, "G": 14.63}
DEFAULT_DATA_COL_WIDTH = 12.6328125


def _cell(ws, row, col_letter, value=None, font=None, fill=None, align=None, number_format=None):
    c = ws.cell(row=row, column=column_index_from_string(col_letter))
    if value is not None:
        c.value = value
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if number_format:
        c.number_format = number_format
    c.border = BORDER_ALL
    return c


def build(order, order_date, rows, out_path):
    columns = order.COLUMNS
    last_col = columns[-1]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = order.DATE_TEXT

    # ширины столбцов - одинаковая логика для ВСЕХ столбцов данных, без
    # индивидуального спец-кейса для последних (это была часть жалобы: почему
    # последние столбцы выглядят иначе - в этой генерации все столбцы строятся
    # одним циклом). Столбцы данных без индивидуальной ширины получают ту же
    # "ширину по умолчанию", что и в эталонном файле, а не более узкий дефолт
    # Excel/openpyxl - иначе они визуально уже, чем в примере.
    for letter, width in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width
    for col in columns:
        if col not in COL_WIDTHS:
            ws.column_dimensions[col].width = DEFAULT_DATA_COL_WIDTH

    # Row 1: полоса упаковки (D - отдельно белая; премиум-столбцы серые; остальные белые)
    _cell(ws, 1, "A", fill=FILL_WHITE)
    _cell(ws, 1, "C", fill=FILL_WHITE)
    for col in columns:
        fill = FILL_PREMIUM_HEADER if col in order.PREMIUM_COLS else FILL_WHITE
        _cell(ws, 1, col, order.HEADER_BAND[col], FONT_HEADER, fill, ALIGN_CENTER)
    _cell(ws, 1, "B", order_date, FONT_TITLE, FILL_WHITE, ALIGN_CENTER, number_format=r"dddd\ dd\.mm\.yy")
    ws.row_dimensions[1].height = 32.25

    # Row 2: заголовок исключения
    _cell(ws, 2, "A", fill=FILL_WHITE)
    _cell(ws, 2, "B", "ИСКЛЮЧЕНИЯ", FONT_TITLE, FILL_WHITE, ALIGN_CENTER)
    _cell(ws, 2, "C", fill=FILL_WHITE)
    for col in columns:
        fill = FILL_PREMIUM_HEADER if col in order.PREMIUM_COLS else FILL_WHITE
        _cell(ws, 2, col, order.EXCEPTION_LABEL[col], FONT_HEADER, fill, ALIGN_CENTER)
    ws.row_dimensions[2].height = 120

    # Row 3: номера столов - ОДИНАКОВАЯ голубая заливка на ВСЕХ столбцах без
    # исключения - единым циклом, а не по группам премиум/стандарт.
    _cell(ws, 3, "A", fill=FILL_ROW3_BLUE)
    _cell(ws, 3, "B", "Номера столов", Font(name="Nunito", size=13, bold=True), FILL_ROW3_BLUE, ALIGN_CENTER)
    _cell(ws, 3, "C", fill=FILL_ROW3_BLUE)
    for col in columns:
        _cell(ws, 3, col, order.TABLE_NUMBER[col], FONT_ROW3, FILL_ROW3_BLUE, ALIGN_CENTER)
    ws.row_dimensions[3].height = 46.5

    # Row 4: количества
    _cell(ws, 4, "A", fill=FILL_WHITE)
    _cell(ws, 4, "B", fill=FILL_WHITE)
    _cell(ws, 4, "C", f"=SUM(D4:{last_col}4)", FONT_ROW4, FILL_WHITE, ALIGN_CENTER)
    for col in columns:
        _cell(ws, 4, col, order.COUNTS.get(col), FONT_ROW4, FILL_WHITE, ALIGN_CENTER)
    ws.row_dimensions[4].height = 36.75

    r = 5
    current_meal = None
    for item in rows:
        if item["meal"] != current_meal:
            current_meal = item["meal"]
            _cell(ws, r, "A", fill=FILL_MEAL_PINK)
            _cell(ws, r, "B", current_meal, FONT_MEAL, FILL_MEAL_PINK, ALIGN_LEFT)
            _cell(ws, r, "C", fill=FILL_MEAL_PINK)
            for col in columns:
                _cell(ws, r, col, order.TABLE_NUMBER[col], FONT_ROW3, FILL_MEAL_PINK, ALIGN_CENTER)
            ws.row_dimensions[r].height = 29.25
            r += 1

        fill = FILL_BASE_GREY if item["fill"] == "base" else FILL_WHITE
        _cell(ws, r, "A", fill=fill)
        _cell(ws, r, "B", item["name"], FONT_DISH_NAME, fill, ALIGN_LEFT)
        _cell(ws, r, "C", f"=SUM(D{r}:{last_col}{r})", FONT_DISH_VAL, fill, ALIGN_CENTER)
        ws.row_dimensions[r].height = 37.5
        for col in columns:
            if col in item["cols"]:
                _cell(ws, r, col, f"={col}4", FONT_DISH_VAL, fill, ALIGN_CENTER)
            else:
                _cell(ws, r, col, None, FONT_DISH_VAL, fill, ALIGN_CENTER)
        r += 1

    ws.freeze_panes = None
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main(order_module="order_grid", out_path=None):
    order = importlib.import_module(order_module)
    out_path = Path(out_path) if out_path else OUTPUT_DIR / f"{order.OUTPUT_NAME}.xlsx"
    rows = order.build_rows()
    path = build(order, order.ORDER_DATE, rows, out_path)
    print(f"Сохранено: {path}")


if __name__ == "__main__":
    main(
        sys.argv[1] if len(sys.argv) > 1 else "order_grid",
        sys.argv[2] if len(sys.argv) > 2 else None,
    )
